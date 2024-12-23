#!/usr/bin/env python
# -*- coding:utf-8 -*-
import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time

def scrape_statista_pages(start_page, end_page, delay=5):
    """
    Scrape Statista pages from start_page to end_page.

    Parameters:
    - start_page (int): The first page number to scrape.
    - end_page (int): The last page number to scrape.
    - delay (int): Delay in seconds between requests to avoid being blocked.
    """
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    }

    all_data = {
        'Topics': [],
        'Report Title': [],
        'Description': [],
        'Published Date': [],
        'Image Source': [],
    }

    for page_number in range(start_page, end_page + 1):
        url = f"https://www.statista.com/chartoftheday/ALL/p/{page_number}/"
        page_text = requests.get(url=url, headers=headers).text

        tree = etree.HTML(page_text)
        topics = tree.xpath('//figure/figcaption/div/h3/text()')
        reports_titles = tree.xpath('//figure/img/@title')
        descriptions = tree.xpath('//figure/figcaption/text()')
        published_dates = tree.xpath('//figure/figcaption/div/time/text()')
        img_src = tree.xpath('//figure/img/@src')

        all_data['Topics'].extend(topics)
        all_data['Report Title'].extend(reports_titles)
        all_data['Description'].extend(descriptions)
        all_data['Published Date'].extend(published_dates)
        all_data['Image Source'].extend(img_src)

        print(f"Page {page_number} scraped. Waiting {delay} seconds before next request...")
        time.sleep(delay)  # Delay to avoid being blocked

    return all_data

def upload_data_to_excel_openpyxl(data, excel_filename):
    wb = Workbook()
    ws = wb.active

    max_length = {}
    for col, header in enumerate(data.keys(), start=1):
        ws.cell(row=1, column=col, value=header)
        max_length[col] = len(header)

    for col_index, (key, values) in enumerate(data.items(), start=1):
        for row_index, value in enumerate(values, start=2):
            ws.cell(row=row_index, column=col_index, value=value)
            max_length[col_index] = max(max_length[col_index], len(str(value)))

    for col, length in max_length.items():
        ws.column_dimensions[get_column_letter(col)].width = length

    wb.save(filename=excel_filename)
    print(f"Data successfully saved to {excel_filename}")

if __name__ == "__main__":
    # Scrape pages 2 through 137 with a delay of 5 seconds between each request
    scraped_data = scrape_statista_pages(2, 137, delay=5)

    # Save scraped data to Excel
    excel_filename = '/Users/xuchenlong/Desktop/Python_Web_Scrapings/Stats/Statista_Data.xlsx'
    upload_data_to_excel_openpyxl(scraped_data, excel_filename)





